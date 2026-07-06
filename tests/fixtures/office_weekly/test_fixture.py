from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

FIXTURE_DIR = Path(__file__).resolve().parent


def _metric_values() -> dict[str, object]:
    payload = json.loads((FIXTURE_DIR / "expected_metrics.json").read_text())
    return {metric["id"]: metric["value"] for metric in payload["metrics"]}


def test_sales_workbook_matches_expected_metrics() -> None:
    workbook = load_workbook(FIXTURE_DIR / "sales_data.xlsx", data_only=True)
    sheet = workbook["sales_data"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert len(rows) == 8

    metrics = _metric_values()
    total_visitors = sum(row[3] for row in rows)
    total_orders = sum(row[4] for row in rows)
    total_gmv = sum(row[5] for row in rows)
    conversion_rate = total_orders / total_visitors

    assert total_gmv == metrics["total_gmv_cny"]
    assert total_orders == metrics["total_orders"]
    assert total_visitors == metrics["total_visitors"]
    assert abs(conversion_rate - metrics["conversion_rate"]) < 1e-10


def test_weekly_fixture_constraints_cover_office_outputs() -> None:
    constraints = json.loads((FIXTURE_DIR / "expected_constraints.json").read_text())
    notes = (FIXTURE_DIR / "meeting_notes.md").read_text()

    assert constraints["outputs"]["pptx_max_pages"] == 6
    assert constraints["numeric_grounding"]["require_fact_ids"] is True
    assert set(constraints["outputs"]["required_artifacts"]) == {
        "verified_facts.json",
        "weekly_report.docx",
        "weekly_review.pptx",
        "quality_report.json",
    }

    for section in constraints["meeting_coverage"]["required_sections"]:
        assert f"## {section}" in notes
