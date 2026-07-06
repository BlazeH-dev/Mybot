"""Generate the deterministic office weekly sales fixture."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

FIXTURE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = FIXTURE_DIR / "sales_data.xlsx"

HEADERS = ["date", "region", "channel", "visitors", "orders", "gmv_cny"]
ROWS = [
    ("2026-06-24", "East", "web", 1200, 96, 288000),
    ("2026-06-24", "South", "app", 900, 72, 216000),
    ("2026-06-25", "North", "web", 650, 39, 117000),
    ("2026-06-25", "East", "app", 1100, 99, 297000),
    ("2026-06-26", "West", "web", 700, 42, 126000),
    ("2026-06-26", "South", "web", 950, 76, 228000),
    ("2026-06-27", "North", "app", 800, 56, 168000),
    ("2026-06-27", "East", "web", 1000, 90, 270000),
]


def build_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sales_data"
    sheet.append(HEADERS)

    for row in ROWS:
        sheet.append(row)

    table = Table(displayName="WeeklySales", ref=f"A1:F{len(ROWS) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 12,
        "E": 10,
        "F": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for cell in sheet["F"][1:]:
        cell.number_format = '#,##0'

    sheet.freeze_panes = "A2"
    return workbook


def main() -> None:
    workbook = build_workbook()
    workbook.save(WORKBOOK_PATH)

    loaded = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = loaded["sales_data"]
    if sheet.max_row != len(ROWS) + 1 or sheet.max_column != len(HEADERS):
        raise RuntimeError("Generated workbook shape does not match expected fixture.")

    print(f"Generated {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
