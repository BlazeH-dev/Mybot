"""Extract deterministic verified facts from an Excel workbook."""

from __future__ import annotations

import argparse
import math
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from nanobot.skills._shared.office_core.common import json_ready, read_json, write_json


def _format_value(value: Any, value_format: str | None) -> str:
    if value_format == "currency_cny":
        return f"CNY {float(value):,.0f}"
    if value_format == "integer":
        return f"{int(value):,}"
    if value_format == "percent":
        return f"{float(value) * 100:.2f}%"
    return str(value)


def _coerce_number(value: Any, *, column: str) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        raise ValueError(f"column {column} contains a boolean where a number is required")
    if isinstance(value, int | float):
        number = float(value)
        if math.isnan(number):
            return 0.0
        return number
    raise ValueError(f"column {column} contains non-numeric value {value!r}")


def _load_rows(input_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if any(not header for header in headers):
        raise ValueError("workbook header row contains empty column names")

    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell is not None for cell in row):
            continue
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
        }
        records.append(record)
    return records


def _require_columns(rows: list[dict[str, Any]], columns: set[str]) -> None:
    available = set(rows[0]) if rows else set()
    missing = sorted(columns - available)
    if missing:
        raise ValueError(f"missing required column(s): {', '.join(missing)}")


def _sum_column(rows: list[dict[str, Any]], column: str) -> float:
    return sum(_coerce_number(row.get(column), column=column) for row in rows)


def _ratio(rows: list[dict[str, Any]], numerator_column: str, denominator_column: str) -> float:
    numerator = _sum_column(rows, numerator_column)
    denominator = _sum_column(rows, denominator_column)
    if denominator == 0:
        raise ValueError(f"denominator column {denominator_column} sums to zero")
    return numerator / denominator


def _top_by_sum(rows: list[dict[str, Any]], group_by: str, value_column: str) -> tuple[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for row in rows:
        group = row.get(group_by)
        if group is None or group == "":
            continue
        totals[str(group)] += _coerce_number(row.get(value_column), column=value_column)
    if not totals:
        raise ValueError(f"no groups found for {group_by}")
    return max(totals.items(), key=lambda item: item[1])


def _fact(
    metric: dict[str, Any],
    value: Any,
    *,
    display_value: str,
    source_file: Path,
    source_columns: list[str],
    row_count: int,
) -> dict[str, Any]:
    fact_id = metric.get("fact_id") or metric.get("id")
    if not isinstance(fact_id, str) or not fact_id:
        raise ValueError("each metric must include fact_id")
    name = metric.get("name") or metric.get("label") or fact_id
    calculation = metric.get("calculation")
    return {
        "fact_id": fact_id,
        "name": str(name),
        "value": json_ready(value),
        "display_value": display_value,
        "unit": metric.get("unit", ""),
        "source": {
            "workbook": source_file.name,
            "sheet": metric.get("sheet"),
            "columns": source_columns,
            "rows": row_count,
        },
        "calculation": str(calculation),
        "confidence": 1.0,
    }


def _validate_facts_payload(payload: dict[str, Any]) -> None:
    required = {
        "fact_id",
        "name",
        "value",
        "display_value",
        "unit",
        "source",
        "calculation",
        "confidence",
    }
    seen: set[str] = set()
    for fact in payload.get("facts", []):
        missing = required - set(fact)
        if missing:
            raise ValueError(f"fact missing required field(s): {', '.join(sorted(missing))}")
        fact_id = fact["fact_id"]
        if fact_id in seen:
            raise ValueError(f"duplicate fact_id: {fact_id}")
        seen.add(fact_id)
        confidence = fact["confidence"]
        if not isinstance(confidence, int | float) or not 0 <= confidence <= 1:
            raise ValueError(f"invalid confidence for {fact_id}: {confidence!r}")


def extract_facts(input_path: Path, spec_path: Path) -> dict[str, Any]:
    spec = read_json(spec_path)
    default_sheet = spec.get("sheet")
    if not isinstance(default_sheet, str) or not default_sheet:
        raise ValueError("metric spec must include a default sheet")

    metrics = spec.get("metrics")
    if not isinstance(metrics, list) or not metrics:
        raise ValueError("metric spec must include a non-empty metrics list")

    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
    facts: list[dict[str, Any]] = []
    for metric in metrics:
        if not isinstance(metric, dict):
            raise ValueError("each metric must be an object")
        sheet = metric.get("sheet") or default_sheet
        if not isinstance(sheet, str) or not sheet:
            raise ValueError("metric sheet must be a string")
        if sheet not in rows_by_sheet:
            rows_by_sheet[sheet] = _load_rows(input_path, sheet)
        rows = rows_by_sheet[sheet]
        metric["sheet"] = sheet

        calculation = metric.get("calculation")
        value_format = metric.get("format")
        if calculation == "sum":
            column = str(metric["column"])
            _require_columns(rows, {column})
            value = _sum_column(rows, column)
            if value.is_integer():
                value = int(value)
            display_value = _format_value(value, value_format)
            facts.append(
                _fact(
                    metric,
                    value,
                    display_value=display_value,
                    source_file=input_path,
                    source_columns=[column],
                    row_count=len(rows),
                )
            )
            continue

        if calculation == "ratio":
            numerator_column = str(metric["numerator_column"])
            denominator_column = str(metric["denominator_column"])
            _require_columns(rows, {numerator_column, denominator_column})
            value = _ratio(rows, numerator_column, denominator_column)
            if value.is_integer():
                value = int(value)
            display_value = _format_value(value, value_format)
            facts.append(
                _fact(
                    metric,
                    value,
                    display_value=display_value,
                    source_file=input_path,
                    source_columns=[numerator_column, denominator_column],
                    row_count=len(rows),
                )
            )
            continue

        if calculation == "top_by_sum":
            group_by = str(metric["group_by"])
            value_column = str(metric["value_column"])
            _require_columns(rows, {group_by, value_column})
            top_label, top_value = _top_by_sum(rows, group_by, value_column)
            value = top_label
            display_total = _format_value(top_value, metric.get("value_format"))
            display_value = f"{top_label} ({display_total})"
            facts.append(
                _fact(
                    metric,
                    value,
                    display_value=display_value,
                    source_file=input_path,
                    source_columns=[group_by, value_column],
                    row_count=len(rows),
                )
            )
            continue

        raise ValueError(f"unsupported calculation: {calculation!r}")

    payload = {
        "schema_version": 1,
        "source_file": input_path.name,
        "facts": facts,
    }
    _validate_facts_payload(payload)
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--in", dest="input_path", required=True, type=Path)
    parser.add_argument("--spec", dest="spec_path", required=True, type=Path)
    parser.add_argument("--out", dest="output_path", required=True, type=Path)
    args = parser.parse_args()

    write_json(args.output_path, extract_facts(args.input_path, args.spec_path))


if __name__ == "__main__":
    main()
