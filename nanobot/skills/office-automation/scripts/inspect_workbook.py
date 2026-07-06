"""Inspect an Excel workbook and emit a compact JSON schema summary."""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from _common import json_ready, write_json
from openpyxl import load_workbook


def _column_name(value: Any, index: int) -> str:
    if value is None or str(value).strip() == "":
        return f"column_{index}"
    return str(value).strip()


def _value_type(value: Any) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int | float):
        return "number"
    if isinstance(value, datetime | date):
        return "date"
    return "text"


def _infer_column_type(values: list[Any]) -> str:
    non_empty = [value for value in values if value is not None and value != ""]
    if not non_empty:
        return "empty"
    counts = Counter(_value_type(value) for value in non_empty)
    if len(counts) == 1:
        return next(iter(counts))
    return "mixed"


def inspect_workbook(input_path: Path, sample_rows: int = 5) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            sheets.append(
                {
                    "name": worksheet.title,
                    "row_count": 0,
                    "column_count": 0,
                    "columns": [],
                }
            )
            continue

        headers = [_column_name(value, index) for index, value in enumerate(rows[0], start=1)]
        data_rows = [row for row in rows[1:] if any(cell is not None for cell in row)]
        columns: list[dict[str, Any]] = []
        for index, header in enumerate(headers):
            values = [row[index] if index < len(row) else None for row in data_rows]
            sample_values = [json_ready(value) for value in values[:sample_rows]]
            columns.append(
                {
                    "name": header,
                    "index": index + 1,
                    "type": _infer_column_type(values[: max(sample_rows, 25)]),
                    "sample_values": sample_values,
                }
            )

        sheets.append(
            {
                "name": worksheet.title,
                "row_count": len(data_rows),
                "column_count": len(headers),
                "columns": columns,
            }
        )

    return {
        "schema_version": 1,
        "source_file": input_path.name,
        "sheets": sheets,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--in", dest="input_path", required=True, type=Path)
    parser.add_argument("--out", dest="output_path", required=True, type=Path)
    parser.add_argument("--sample-rows", type=int, default=5)
    args = parser.parse_args()

    payload = inspect_workbook(args.input_path, sample_rows=args.sample_rows)
    write_json(args.output_path, payload)


if __name__ == "__main__":
    main()
